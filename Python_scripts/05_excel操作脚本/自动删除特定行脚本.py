#!/usr/bin/env python3
"""
Excel行删除脚本：扫描第一列，包含特定字符串则删除整行
使用方法: python delete_rows.py <输入文件> <目标字符串> [输出文件]
"""

import sys
import openpyxl
from openpyxl.utils import column_index_from_string


def delete_rows_by_first_column(input_file, target_string, output_file=None, case_sensitive=False):
    """
    删除Excel文件中第一列包含特定字符串的行

    参数:
        input_file: 输入的xlsx文件路径
        target_string: 要匹配的字符串
        output_file: 输出文件路径（默认覆盖原文件）
        case_sensitive: 是否大小写敏感（默认不敏感）
    """
    if output_file is None:
        output_file = input_file

    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active  # 使用活动工作表

        # 获取总行数
        max_row = ws.max_row

        # 用于记录删除的行数
        deleted_count = 0

        # 从最后一行开始向前遍历（避免删除后行号变化）
        for row_idx in range(max_row, 0, -1):
            # 获取第一列的单元格值（列索引为1）
            cell_value = ws.cell(row=row_idx, column=1).value

            # 检查单元格是否不为空且包含目标字符串
            if cell_value is not None:
                # 转换为字符串处理
                cell_text = str(cell_value)

                # 根据大小写敏感设置进行匹配
                if case_sensitive:
                    match = target_string in cell_text
                else:
                    match = target_string.lower() in cell_text.lower()

                # 如果匹配，删除整行
                if match:
                    ws.delete_rows(row_idx)
                    deleted_count += 1

        # 保存修改后的工作簿
        wb.save(output_file)
        print(f"✅ 处理完成！共删除 {deleted_count} 行")
        print(f"💾 结果已保存至: {output_file}")

    except FileNotFoundError:
        print(f"❌ 错误: 文件 '{input_file}' 不存在")
        sys.exit(1)
    except Exception as e:
        print(f"❌ 错误: {str(e)}")
        sys.exit(1)


def main():
    """命令行入口"""
    if len(sys.argv) < 3:
        print("使用方法: python delete_rows.py <输入文件> <目标字符串> [输出文件]")
        print("示例: python delete_rows.py data.xlsx '删除标记' output.xlsx")
        sys.exit(1)

    input_file = sys.argv[1]
    target_string = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) > 3 else None

    # 执行删除操作（默认不区分大小写）
    delete_rows_by_first_column(input_file, target_string, output_file, case_sensitive=False)


if __name__ == "__main__":
    main()